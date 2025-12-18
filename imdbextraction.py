"""
IMDB scraper - Automated based scraper,
multi-page scraper, handles same HTML
structure of website for many links. Dec, 2025
"""

# ==============================================================
# PACKAGES
# ==============================================================

from playwright.sync_api import sync_playwright
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

# ===============================================================
# SCRAPING FUNCTION
# ===============================================================

def scraping_process():
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        page = browser.new_page()
        page.goto("https://www.imdb.com/chart/top/")
        print("""
        |===============================================|
        |                                               |
        |  - We are opened chromium browser, currently  |
        |  using provided selectors to get the data.    |
        |  - It may take 10 upto 30 seconds to complete |
        |  the process.                                 |
        |  - Please wait for output or Excel file from  |
        |  project root.                                |
        |                                               |
        |===============================================|
        """)

        page.wait_for_selector(".ipc-metadata-list-summary-item", timeout=10000)
        movie_containers = page.query_selector_all(".ipc-metadata-list-summary-item")

        titles = []
        years = []
        rates = []
        lengths = []

        for container in movie_containers:
            title_elem = container.query_selector(".ipc-title-link-wrapper")
            title = title_elem.text_content().strip() if title_elem else "Not found"
            titles.append(title)

            year_elem = container.query_selector(".cli-title-metadata-item")
            year = year_elem.text_content().strip() if year_elem else "Not found"
            years.append(year)

            rating_elem = container.query_selector(".ipc-rating-star--rating")
            rating = rating_elem.text_content().strip() if rating_elem else "Not found"
            rates.append(rating)

            length_elem = container.query_selector_all(".sc-b4f120f6-7.hoOxkw.cli-title-metadata-item")
            length = length_elem[1].text_content().strip() if len(length_elem) > 1 else "Not found"
            lengths.append(length)
            

        browser.close()
        return titles, years, rates, lengths

# ================================================================
# SAVING PROCESS
# ================================================================

def saving_process(movie_titles, years, rates, lengths):
    df = pd.DataFrame({"Title": movie_titles, "Year": years, "Rating": rates, "Length": lengths})
    excel_file = "Movie_Data_From_Imdb.xlsx"
    df.to_excel(excel_file, index=False)
    wb = load_workbook(excel_file)
    ws = wb.active
    max_title_len = max((len(str(title)) for title in movie_titles), default=10)
    max_year_len = max((len(str(year)) for year in years), default=4)
    max_rating_len = max((len(str(rate)) for rate in rates), default=3)
    max_length_len = max((len(str(length)) for length in lengths), default = 0)
    ws.column_dimensions[get_column_letter(1)].width = max_title_len + 2
    ws.column_dimensions[get_column_letter(2)].width = max_year_len + 2
    ws.column_dimensions[get_column_letter(3)].width = max_rating_len + 2
    ws.column_dimensions[get_column_letter(4)].width = max_length_len + 2
    wb.save(excel_file)

# =================================================================
# ENTRY POINT
# =================================================================

if __name__ == "__main__":
    titles, years, rates, lengths = scraping_process()
    saving_process(titles, years, rates, lengths)
    print("Success!")