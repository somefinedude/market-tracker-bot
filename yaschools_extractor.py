import random
import time
import datetime
import pandas as lk
from playwright.sync_api import sync_playwright
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

sleep_dur = random.randint(1, 7)

def human_sleep(min=0.5, max=4.7):
    time.sleep(random.uniform(min, max))

def extraction_proc():
    with sync_playwright() as p:
        chrome = p.chromium.launch(headless=False)
        context = chrome.new_context(
        user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120 Safari/537.36",
        viewport={"width": 1570, "height": 900},
        locale="en-US"
        )
        page = context.new_page()
        try:
            page.goto("https://yaschools.com/en", wait_until="load", timeout=0)
        except Exception as raised_error:
            print("Failed to lead the page", raised_error)

        print("""
        |========================================================|
        |                                                        |
        |  - Getting school informations from 'yaschools.com'    |
        |  - Automated scraping proccess, currently using pro    |
        |    vided selectors and elements.                       |
        |  - Results will be delivered in Excel file, wait for   |
        |    ouput file or message from terminal.                |
        |  - We use human-immitating patterns, so please be pa   |
        |    tient                                               |
        |                                                        |
        |========================================================|
        """)

        page.wait_for_load_state("load")
        human_sleep(1.8, 3.3)
        page.wait_for_selector('a[href="/en/schools"]')
        page.click('a[href="/en/schools"]')
        page.wait_for_load_state("load")
        human_sleep(1.7, 4.1)

        data = []

        while True:
            school_cards = page.query_selector_all("div.school")
            for card in school_cards:
                try:
                    name = card.query_selector("p a").inner_text().strip()
                except:
                    name = "Null"

                try:
                    logo = card.query_selector(".top .img a img").get_attribute("src")
                except:
                    logo = "Null"

                try:
                    rating = card.query_selector(".ratebadge").inner_text().strip()
                except:
                    rating = "Null"

                try:  
                    comments = card.query_selector("p[style*='font-size: 12px']").inner_text().strip()
                except:
                    comments = "Null"
                    
                try:
                    school_type = card.query_selector("div p[style*='background']").inner_text().strip()
                except:
                    school_type = "Null"
                
                try:
                    location = card.query_selector("p.city").inner_text().strip()
                except:
                    location = "Null"

                try:   
                    fees = card.query_selector(".feesText div").inner_text().strip()
                except:
                    fees = "Null"

                try:    
                    details_url = card.query_selector("a[href*='/en/school/']").get_attribute("href")
                except:
                    details_url = "Null"

                data.append({
                    "Name": name,
                    "Logo": logo,
                    "Rating": rating,
                    "Comments": comments,
                    "School type": school_type,
                    "Location": location,
                    "Starting fee": fees,
                    "URL": details_url
                })
            try:
                next_btn = page.query_selector("li.next a")
                if next_btn:
                    next_url = next_btn.get_attribute("href")
                    page.goto("https://yaschools.com" + next_url, wait_until="load")
                    human_sleep(1.5, 3.5)
                else:
                    break
            except:
                break


        df = lk.DataFrame(data)
        excel_file = "Schools_In_UAE.xlsx"
        df.to_excel(excel_file, index=False)
        wb = load_workbook(excel_file)
        ws = wb.active
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max_length + 2
        wb.save(excel_file)
        print("Success!")
extraction_proc()